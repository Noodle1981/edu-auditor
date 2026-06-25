import openpyxl
import sqlite3
import os

db_path = r"database/database.sqlite"
excel_path = r"tabla_edificios_radios.xlsx"

if not os.path.exists(excel_path):
    print(f"Error: Excel file not found at {excel_path}!")
    exit(1)

if not os.path.exists(db_path):
    print(f"Error: SQLite database not found at {db_path}!")
    exit(1)

print("Starting import process...")
print(f"Loading Excel file: {excel_path}...")
wb = openpyxl.load_workbook(excel_path, data_only=True)
sheet = wb.active
print(f"Active sheet: {sheet.title} | Total rows: {sheet.max_row}")

print(f"Connecting to database: {db_path}...")
conn = sqlite3.connect(db_path)
cursor = conn.cursor()

# Disable foreign keys temporarily to do fast updates
cursor.execute("PRAGMA foreign_keys = OFF;")

stats = {
    'processed_rows': 0,
    'ignored_private': 0,
    'buildings_matched': 0,
    'buildings_updated': 0,
    'modalities_updated': 0,
}

for r in range(2, sheet.max_row + 1):
    cui_raw = sheet.cell(row=r, column=3).value # Column 3 is CUI
    ambito_raw = sheet.cell(row=r, column=17).value # Column 17 is ambito
    
    if cui_raw is None:
        continue
        
    stats['processed_rows'] += 1
    
    # Filter only PUBLIC
    ambito_str = str(ambito_raw).strip().upper() if ambito_raw is not None else ""
    if "PUBLIC" not in ambito_str:
        stats['ignored_private'] += 1
        continue
        
    cui_str = str(cui_raw).strip()
    
    # Extract other metrics safely
    punto_partida = sheet.cell(row=r, column=8).value
    dist_circunf = sheet.cell(row=r, column=9).value
    radio_circ = sheet.cell(row=r, column=10).value
    distancia_camino = sheet.cell(row=r, column=11).value
    radio_camino = sheet.cell(row=r, column=12).value
    tiempo_google_auto = sheet.cell(row=r, column=13).value
    radio_sige = sheet.cell(row=r, column=14).value
    observacion = sheet.cell(row=r, column=18).value
    lat_excel = sheet.cell(row=r, column=6).value
    lon_excel = sheet.cell(row=r, column=5).value
    
    # Normalize tiempo_google_auto if it is a datetime.time object
    if hasattr(tiempo_google_auto, 'strftime'):
        tiempo_google_auto = tiempo_google_auto.strftime('%H:%M:%S')
    elif tiempo_google_auto is not None:
        tiempo_google_auto = str(tiempo_google_auto).strip()
        
    # Convert numerical values safely
    try:
        dist_circunf = float(dist_circunf) if dist_circunf is not None else None
    except ValueError:
        dist_circunf = None
        
    try:
        radio_circ = int(radio_circ) if radio_circ is not None else None
    except ValueError:
        radio_circ = None
        
    try:
        distancia_camino = float(distancia_camino) if distancia_camino is not None else None
    except ValueError:
        distancia_camino = None
        
    try:
        radio_camino = int(radio_camino) if radio_camino is not None else None
    except ValueError:
        radio_camino = None
        
    try:
        radio_sige = int(radio_sige) if radio_sige is not None else None
    except ValueError:
        radio_sige = None

    try:
        lat_excel = float(lat_excel) if lat_excel is not None else None
    except ValueError:
        lat_excel = None

    try:
        lon_excel = float(lon_excel) if lon_excel is not None else None
    except ValueError:
        lon_excel = None
        
    # Search for building matching CUI
    # cui in DB is an integer column, so try to convert CUI to integer, or search directly
    try:
        cui_int = int(cui_str)
        cursor.execute("SELECT id FROM edificios WHERE cui = ?;", (cui_int,))
    except ValueError:
        cursor.execute("SELECT id FROM edificios WHERE cui = ?;", (cui_str,))
        
    building_row = cursor.fetchone()
    if building_row:
        building_id = building_row[0]
        stats['buildings_matched'] += 1
        
        # Update building
        cursor.execute("""
            UPDATE edificios 
            SET punto_partida = ?, dist_circunf = ?, radio_circ = ?, 
                distancia_camino = ?, radio_camino = ?, tiempo_google_auto = ?, 
                observacion = ?, latitud = ?, longitud = ?, updated_at = CURRENT_TIMESTAMP
            WHERE id = ?;
        """, (punto_partida, dist_circunf, radio_circ, distancia_camino, radio_camino, tiempo_google_auto, observacion, lat_excel, lon_excel, building_id))
        
        if cursor.rowcount > 0:
            stats['buildings_updated'] += 1
            
        # If we have radio_sige, update modalities of this building
        if radio_sige is not None:
            # Find all establishments belonging to this building
            cursor.execute("SELECT id FROM establecimientos WHERE edificio_id = ?;", (building_id,))
            est_rows = cursor.fetchall()
            for est_row in est_rows:
                est_id = est_row[0]
                # Update modalities where establishment_id matches and ambito is PUBLICO
                cursor.execute("""
                    UPDATE modalidades
                    SET radio_sige = ?, updated_at = CURRENT_TIMESTAMP
                    WHERE establecimiento_id = ? AND ambito = 'PUBLICO';
                """, (radio_sige, est_id))
                stats['modalities_updated'] += cursor.rowcount

# Re-enable foreign keys
cursor.execute("PRAGMA foreign_keys = ON;")
conn.commit()
conn.close()

print("\nImport process completed successfully!")
print("Statistics:")
for k, v in stats.items():
    print(f" - {k}: {v}")
